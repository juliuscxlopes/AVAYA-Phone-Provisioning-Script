import openpyxl
import os
import tkinter as tk
from tkinter import scrolledtext
from tkinter import simpledialog
from pathlib import Path

class Elementos_graficos:
    def __init__(self, root):
        self.root = root
        self.root.title("Status de Processamento")

        self.output_text = scrolledtext.ScrolledText(self.root, wrap=tk.WORD)
        self.output_text.pack(fill=tk.BOTH, expand=True)

        self.botao_processar = tk.Button(self.root, text="Gerar Arquivos", command=self.processar_arquivos)
        self.botao_processar.pack()

    def processar_arquivos(self):
        diretorio_base = Path(os.path.expanduser('~')) / 'Projetos' / 'Importação_AVAYA'
        arquivos_xlsx = list(diretorio_base.glob('*.xlsx'))

        if not arquivos_xlsx:
            self.output_text.insert(tk.END, "Nenhum arquivo XLSX encontrado no diretório base.\n")
        else:
            caminho_padrao = diretorio_base / 'Padrão.txt'
            linhas_padrao = self.ler_linhas_padrao(caminho_padrao)

            for caminho_planilha in arquivos_xlsx:
                nome_planilha = os.path.splitext(caminho_planilha.name)[0]
                diretorio_saida = diretorio_base / nome_planilha

                if not diretorio_saida.exists():
                    diretorio_saida.mkdir(parents=True)

                planilha = openpyxl.load_workbook(caminho_planilha)
                folha = planilha.active
                
                num_linhas = folha.max_row  # Obtém o número total de linhas na planilha
                titulos_colunas = [cell.value.lower() if cell.value else '' for cell in folha[1]]  # Converta para minúsculas
                
                if "ramal" in titulos_colunas and "mac" in titulos_colunas:  # Verifica se as colunas "ramal" e "mac" estão presentes
                    arquivos_gerados = self.processar_planilha(planilha, diretorio_saida, linhas_padrao, nome_planilha)
                    if arquivos_gerados:   
                        self.output_text.insert(tk.END, f"Todos os ramais da \"{nome_planilha}\" foram criados e salvos:\n")
                        for arquivo in arquivos_gerados:
                            self.output_text.insert(tk.END, f"- {arquivo}\n")
                        self.output_text.insert(tk.END, f"{num_linhas} Arquivos gerados com sucesso!!\n\n")
                else:
                    self.output_text.insert(tk.END, f"Não foi encontrado o título 'RAMAL' ou 'MAC' na planilha '{nome_planilha}'.\n")

    def ler_linhas_padrao(self, caminho_padrao):
        with open(caminho_padrao, 'r') as txt:
            return txt.readlines()

    def criar_arquivo(self, nome_arquivo, linhas_modificadas):
        with open(nome_arquivo, 'w') as txt:
            txt.writelines(linhas_modificadas)

    def processar_planilha(self, planilha, diretorio_saida, linhas_padrao, nome_planilha):
        self.output_text.insert(tk.END, f"Processando planilha: {nome_planilha}.\n")
        folha = planilha.active
        arquivos_gerados = []
        
        titulos_colunas = [cell.value.lower() if cell.value is not None else '' for cell in folha[1]]
        ramal_index = titulos_colunas.index("ramal")
        mac_index = titulos_colunas.index("mac")
        
        num_linhas = sum(1 for row in folha.iter_rows(min_row=2) if any(cell.value for cell in row))
        if num_linhas < 1:  # Verifique se há pelo menos uma linha de dados (além do título)
            self.output_text.insert(tk.END, f"A planilha '{nome_planilha}' não tem dados suficientes para processar.\n")
            return  # Sair da função pois não há dados suficientes para processar.

        prefixo = None
        
        for row in folha.iter_rows(min_row=2):
            ramal = str(row[ramal_index].value)
            mac = str(row[mac_index].value)
            
            if len(ramal) != 12:
                if prefixo is None: #Se o prefixo ainda nao foi definido
                    prefixo = tk.simpledialog.askstring("Prefixo", "Insira o prefixo para o ramal:")
                    if prefixo is not None:
                        continue #Se o usuário cancelar a caixa de diálogo, interrompe o sistema.
                
                ramal = prefixo + ramal #concatena o prefixo ao ramal
                
            linhas_modificadas = []
            for linha in linhas_padrao:
                if 'SET FORCE_SIP_USERNAME' in linha:
                    linhas_modificadas.append(f'SET FORCE_SIP_USERNAME {ramal}\n')
                elif 'SET FORCE_SIP_EXTENSION' in linha:
                    linhas_modificadas.append(f'SET FORCE_SIP_EXTENSION {ramal}\n')
                elif 'SET SIP_USER_ACCOUNT' in linha:
                    linhas_modificadas.append(f'SET SIP_USER_ACCOUNT {ramal}@10.159.0.54\n')
                elif 'SET PREV_SIP_USER_ACCOUNT' in linha:
                    linhas_modificadas.append(f'SET PREV_SIP_USER_ACCOUNT {ramal}@10.159.0.54\n')
                elif 'SET DISPLAY_NAME' in linha:
                    linhas_modificadas.append(f'SET DISPLAY_NAME {ramal}\n')
                else:
                    linhas_modificadas.append(linha)
                        
            mac = str(row[mac_index].value).replace(':', '')
            if mac is not None:  # Verifica se 'mac' não é None
                nome_arquivo = f"{mac}.txt"
                caminho_completo = os.path.join(diretorio_saida, nome_arquivo)
                self.criar_arquivo(caminho_completo, linhas_modificadas)
                arquivos_gerados.append(nome_arquivo)
            else:
                pass    

            
        return arquivos_gerados

if __name__ == "__main__":
    root = tk.Tk()
    app = Elementos_graficos(root)
    root.mainloop()
